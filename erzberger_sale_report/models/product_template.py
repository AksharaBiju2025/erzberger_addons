# -*- coding: utf-8 -*-

from odoo import models, fields

import logging
from odoo.osv import expression
_logger = logging.getLogger(__name__)
import base64
from io import BytesIO
from openpyxl import load_workbook


class ProductTemplate(models.Model):
    _inherit = "product.template"

    min_qty_alert = fields.Float(
        string="Minimum Quantity Alert"
    )

    def _cron_product_category_mapping(self):

        attachment = self.env['ir.attachment'].search([
            ('name', '=', 'product_p_categ_maping.xlsx')
        ], limit=1)

        if not attachment:
            _logger.info("Category mapping file not found")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True
        )

        sheet = workbook.active

        updated = 0
        not_found = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            product_name = row[0] and str(row[0]).strip()
            category_name = row[1] and str(row[1]).strip()

            if not product_name or not category_name:
                continue

            product = self.search([
                ('name', '=', product_name)
            ], limit=1)

            category = self.env['product.category'].search([
                ('name', '=', category_name)
            ], limit=1)

            if product and category:
                product.categ_id = category.id
                updated += 1
            else:
                not_found.append(
                    "%s -> %s" % (
                        product_name,
                        category_name
                    )
                )

        _logger.info(
            "Category mapping completed. Updated %s products",
            updated
        )

        for item in not_found:
            _logger.warning(item)


    def _cron_update_unger_products(self):

        company = self.env['res.company'].browse(1)  # replace with actual company id

        category = self.env['product.category'].search([
            ('name', '=', 'Unger')
        ], limit=1)

        uom = self.env['uom.uom'].search([
            ('name', '=', 'Stück')
        ], limit=1)

        if not category:
            return

        if not uom:
            return

        products = self.search([
            ('company_id', '=', company.id)
        ])

        products.write({
            'categ_id': category.id,
            'uom_id': uom.id,
        })

        _logger.info(
            "Updated %s products with category Unger and UoM Stück",
            len(products)
        )



    def _cron_inventory_alert(self):

        brands = [
            'Hubid',
            'Naked',
            'Wendt',
            'Kühn',
            'Coal',
            'Neumann',
            'Ulbricht',
            'Tits',
            'DWU',
            'Wagner',
        ]

        domain = [('min_qty_alert', '>', 0)]

        brand_domain = []
        for brand in brands:
            brand_domain.append(('name', 'ilike', brand))

        products = self.search(
            expression.AND([
                domain,
                expression.OR([[d] for d in brand_domain])
            ])
        )

        template = self.env.ref(
            'erzberger_sale_report.mail_template_inventory_alert',
            raise_if_not_found=False
        )

        if not template:
            return

        inventory_admins = self.env['res.users'].search([
            ('group_ids', 'in', self.env.ref('stock.group_stock_manager').id),
            ('email', '!=', False),
        ])

        for product in products:

            # Check brand name in product name
            if not any(
                    brand.lower() in (product.name or '').lower()
                    for brand in brands
            ):
                continue

            # Check stock level
            if product.qty_available > product.min_qty_alert:
                continue

            for user in inventory_admins:
                _logger.info(
                    "Sending inventory alert for %s to %s",
                    product.name,
                    user.email
                )

                template.send_mail(
                    product.id,
                    email_values={
                        'email_to': user.email,
                    },
                    force_send=True
                )