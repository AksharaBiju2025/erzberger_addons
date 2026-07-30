# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import logging
_logger = logging.getLogger(__name__)
import base64
import io
from io import BytesIO
from openpyxl import load_workbook
BATCH_SIZE = 1000

class ProductTemplate(models.Model):
    _inherit = "product.template"

    # ------------------------------------------------------------------
    # This function about missing product categories will update while trigger this function
    # ------------------------------------------------------------------
    @api.model
    def map_product_categories_from_excel(self):
        _logger.info("========== Product Category Mapping Started ==========")

        attachment = self.env["ir.attachment"].search(
            [("name", "=", "500_products_holz.xlsx")], limit=1
        )

        if not attachment:
            _logger.error("Attachment '500_products_holz.xlsx' not found.")
            return False

        _logger.info("Attachment found: %s", attachment.name)

        import openpyxl
        import io
        import base64

        file_bytes = base64.b64decode(attachment.datas)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        headers = {}
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = idx

        _logger.info("Excel Headers: %s", headers)

        name_col = headers.get("Name")
        default_code_col = headers.get("intere Referenz")
        product_category_col = headers.get("Produktkategorie")
        pos_category_col = headers.get("Kategorie des Kassensystems")

        if not all([
            name_col,
            default_code_col,
            product_category_col,
            pos_category_col,
        ]):
            _logger.error(
                "Required columns missing. "
                "Name=%s, Default Code=%s, Product Category=%s, POS Category=%s",
                name_col,
                default_code_col,
                product_category_col,
                pos_category_col,
            )
            raise UserError("Required columns are missing in the Excel.")

        ProductCategory = self.env["product.category"]
        PosCategory = self.env["pos.category"]

        product_category_cache = {}
        pos_category_cache = {}

        mapped = 0
        not_found = []

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            product_name = str(row[name_col - 1] or "").strip()
            default_code = str(row[default_code_col - 1] or "").strip()
            product_category_name = str(row[product_category_col - 1] or "").strip()
            pos_category_name = str(row[pos_category_col - 1] or "").strip()

            if not product_name:
                continue

            _logger.info(
                "Row %s -> Name='%s', Default Code='%s', Product Category='%s', POS Category='%s'",
                row_no,
                product_name,
                default_code,
                product_category_name,
                pos_category_name,
            )

            domain = [("name", "=", product_name)]
            if default_code:
                domain.append(("default_code", "=", default_code))

            product = self.search(domain, limit=1)

            if not product:
                _logger.warning(
                    "Product not found: Name='%s', Default Code='%s'",
                    product_name,
                    default_code,
                )
                not_found.append((product_name, default_code))
                continue

            _logger.info(
                "Matched Product: %s (ID: %s)",
                product.display_name,
                product.id,
            )

            vals = {}

            # Product Category
            if product_category_name:
                if product_category_name not in product_category_cache:
                    product_category_cache[product_category_name] = ProductCategory.search(
                        [("name", "=", product_category_name)],
                        limit=1,
                    )

                category = product_category_cache[product_category_name]

                if category:
                    vals["categ_id"] = category.id
                    _logger.info(
                        "Mapped Product Category '%s' -> ID %s",
                        category.name,
                        category.id,
                    )
                else:
                    _logger.warning(
                        "Product Category not found: '%s'",
                        product_category_name,
                    )

            # POS Category
            if pos_category_name:
                if pos_category_name not in pos_category_cache:
                    pos_category_cache[pos_category_name] = PosCategory.search(
                        [("name", "=", pos_category_name)],
                        limit=1,
                    )

                pos_category = pos_category_cache[pos_category_name]

                if pos_category:
                    vals["pos_categ_ids"] = [(6, 0, [pos_category.id])]
                    _logger.info(
                        "Mapped POS Category '%s' -> ID %s",
                        pos_category.name,
                        pos_category.id,
                    )
                else:
                    _logger.warning(
                        "POS Category not found: '%s'",
                        pos_category_name,
                    )

            if vals:
                product.write(vals)
                mapped += 1
                _logger.info(
                    "Updated Product '%s' (ID: %s)",
                    product.display_name,
                    product.id,
                )

        _logger.info("========== Product Category Mapping Finished ==========")
        _logger.info("Total Products Updated: %s", mapped)
        _logger.info("Products Not Found: %s", len(not_found))

        if not_found:
            _logger.warning("Not Found Products: %s", not_found)

        return True

    @api.model
    def map_manufacturer_from_excel(self):

        attachment = self.env["ir.attachment"].search(
            [("name", "=", "verpack_man.xlsx")], limit=1
        )
        if not attachment:
            print("Attachment 'verpack_man.xlsx' not found — aborting.")
            return False

        import openpyxl

        file_bytes = base64.b64decode(attachment.datas)
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # Read headers
        headers = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[str(cell.value).strip()] = idx + 1

        name_col = headers.get("Name")
        manufacturer_col = headers.get("Hersteller")
        manufacturer_ref_col = headers.get("Referenz Hersteller")

        if not (name_col and manufacturer_col and manufacturer_ref_col):
            print(
                "Required columns not found.\n"
                f"Beschreibung: {name_col}\n"
                f"Hersteller: {manufacturer_col}\n"
                f"Referenz Hersteller: {manufacturer_ref_col}"
            )
            return False

        # Build lookup
        by_name = {}

        for row in ws.iter_rows(min_row=2, values_only=False):
            product_name = row[name_col - 1].value
            manufacturer = row[manufacturer_col - 1].value
            manufacturer_ref = row[manufacturer_ref_col - 1].value

            if not product_name:
                continue

            by_name[str(product_name).strip()] = {
                "hersteller": (
                    str(manufacturer).strip() if manufacturer else False
                ),
                "referenz_hersteller": (
                    str(manufacturer_ref).strip() if manufacturer_ref else False
                ),
            }

        products = self.search([])
        print(f"Found {len(products)} products")

        matched = 0
        no_excel_row = []
        updated = 0

        for product in products:
            vals = by_name.get((product.name or "").strip())

            if not vals:
                no_excel_row.append(product.id)
                continue

            write_vals = {}

            if vals["hersteller"]:
                write_vals["hersteller"] = vals["hersteller"]

            if vals["referenz_hersteller"]:
                write_vals["referenz_hersteller"] = vals["referenz_hersteller"]

            if write_vals:
                product.write(write_vals)
                updated += 1

            matched += 1

        print(
            f"Manufacturer mapping completed.\n"
            f"Matched: {matched}\n"
            f"Updated: {updated}\n"
            f"No Excel row: {len(no_excel_row)}"
        )

        if no_excel_row:
            print(f"Products with no matching Excel row: {no_excel_row}")

        return True

    @api.model
    def cron_enable_track_inventory(self):
        domain = [
            ("type", "=", "consu"),
            ("is_storable", "=", False),
        ]

        total_updated = 0

        while True:
            products = self.search(domain, limit=BATCH_SIZE)

            if not products:
                break

            products.write({
                "is_storable": True,
            })

            total_updated += len(products)
            self.env.cr.commit()  # Optional for cron jobs with large datasets

            _logger.info(
                "Updated %s products in current batch. Total updated: %s",
                len(products),
                total_updated,
            )

        _logger.info(
            "Track Inventory enabled successfully for %s products.",
            total_updated,
        )

    @api.model
    def cron_enable_dropship_route(self):
        company_id = 2  # Erzberger Verpackung

        dropship_route = self.env.ref(
            "stock_dropshipping.route_drop_shipping",
            raise_if_not_found=False,
        )

        if not dropship_route:
            _logger.warning("Dropship route not found.")
            return

        products = self.search([
            ('company_id', '=', company_id),
            ('route_ids', 'not in', dropship_route.ids),
        ])

        _logger.info("Found %s products.", len(products))

        products.write({
            'route_ids': [(4, dropship_route.id)]
        })

        _logger.info("Dropship route enabled successfully.")

    def _cron_product_category_mapping_verpackungssysteme(self):

        attachment = self.env['ir.attachment'].search([
            ('name', '=', 'product_p_categ_maping _verpackungssysteme.xlsx')
        ], limit=1)

        if not attachment:
            _logger.info("Category mapping file not found")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True
        )

        sheet = workbook.active

        updated = 0
        not_found = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            product_name = str(row[0]).strip() if len(row) > 0 and row[0] else False
            category_name = str(row[1]).strip() if len(row) > 1 and row[1] else False
            purchase_desc = str(row[2]).strip() if len(row) > 2 and row[2] else False

            if not product_name or not category_name:
                continue

            product = self.search([
                ('name', '=', product_name)
            ], limit=1)

            category = self.env['product.category'].search([
                ('name', '=', category_name)
            ], limit=1)
            if product:
                product.description_purchase = purchase_desc

            if product and category:
                product.categ_id = category.id
                updated += 1
            else:
                not_found.append(
                    "%s -> %s" % (
                        product_name,
                        category_name
                    )
                )

        _logger.info(
            "Category mapping completed. Updated %s products",
            updated
        )

        for item in not_found:
            _logger.warning(item)


    def _cron_product_category_mapping_holzkunst(self):

        attachment = self.env['ir.attachment'].search([
            ('name', '=', 'product_p_categ_maping_holzkunst.xlsx')
        ], limit=1)

        if not attachment:
            _logger.info("Category mapping file not found")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True
        )

        sheet = workbook.active

        updated = 0
        not_found = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            product_name = row[0] and str(row[0]).strip()
            category_name = row[1] and str(row[1]).strip()

            if not product_name or not category_name:
                continue

            product = self.search([
                ('name', '=', product_name)
            ], limit=1)

            category = self.env['product.category'].search([
                ('name', '=', category_name)
            ], limit=1)

            if product and category:
                product.categ_id = category.id
                updated += 1
            else:
                not_found.append(
                    "%s -> %s" % (
                        product_name,
                        category_name
                    )
                )

        _logger.info(
            "Category mapping completed. Updated %s products",
            updated
        )

        for item in not_found:
            _logger.warning(item)

    def _cron_quotation_description_mapping(self):

        attachment = self.env['ir.attachment'].search([
            ('name', '=', 'quotation_description_mapping.xlsx')
        ], limit=1)

        if not attachment:
            _logger.info("Quotation description mapping file not found")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True
        )

        sheet = workbook.active

        updated = 0
        not_found = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            product_name = row[0] and str(row[0]).strip()
            quotation_description = row[1] and str(row[1]).strip()

            if not product_name:
                continue

            product = self.search([
                ('name', '=', product_name)
            ], limit=1)

            if product:
                product.description_sale = quotation_description or ""
                updated += 1
            else:
                not_found.append(product_name)

        _logger.info(
            "Quotation description mapping completed. Updated %s products",
            updated
        )

        for item in not_found:
            _logger.warning(
                "Product not found: %s",
                item
            )


    def _cron_set_company1_uom_to_piece(self):

        company = self.env['res.company'].browse(1)
        if not company.exists():
            _logger.info("Company 1 not found.")
            return

        uom_piece = self.env['uom.uom'].search([
            ('name', '=', 'Stück')
        ], limit=1)

        if not uom_piece:
            _logger.info("UoM 'Stück' not found.")
            return

        products = self.search([
            ('company_id', '=', company.id)
        ])

        _logger.info("Updating %s products...", len(products))

        products.write({
            'uom_id': uom_piece.id,
        })

        _logger.info("Finished updating Company 1 products.")

    def _cron_product_uom_weight_mapping(self):

        attachment = self.env['ir.attachment'].search([
            ('name', '=', 'uom_change_ver_cleaned.xlsx')
        ], limit=1)

        if not attachment:
            _logger.info("UoM mapping file not found")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True
        )

        sheet = workbook.active

        updated = 0
        not_found = []
        uom_not_found = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            product_name = str(row[0]).strip() if row[0] else False
            uom_name = str(row[1]).strip() if row[1] else False
            weight = row[2]

            if not product_name:
                continue

            product = self.search([
                ('name', '=', product_name)
            ], limit=1)

            if not product:
                not_found.append(product_name)
                continue

            vals = {}

            # Update UoM
            if uom_name:
                uom = self.env['uom.uom'].search([
                    ('name', '=', uom_name)
                ], limit=1)

                if uom:
                    vals.update({
                        'uom_id': uom.id,
                    })
                else:
                    uom_not_found.append(
                        f"{product_name} -> {uom_name}"
                    )

            # Update Weight
            if weight not in (None, "", False):
                try:
                    vals['weight'] = float(weight)
                except Exception:
                    _logger.warning(
                        "Invalid weight '%s' for product '%s'",
                        weight,
                        product_name
                    )

            if vals:
                product.write(vals)
                updated += 1

        _logger.info(
            "UoM & Weight mapping completed. Updated %s products.",
            updated
        )

        for product_name in not_found:
            _logger.warning(
                "Product not found: %s",
                product_name
            )

        for item in uom_not_found:
            _logger.warning(
                "UoM not found: %s",
                item
            )

    def _cron_product_pos_category_mapping(self):

        attachment = self.env['ir.attachment'].search([
            ('name', '=', 'product_pos_category.xlsx')
        ], limit=1)

        if not attachment:
            _logger.info("POS Category mapping file not found.")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True
        )

        sheet = workbook.active

        updated = 0
        product_not_found = []
        category_not_found = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            product_name = str(row[0]).strip() if row[0] else False
            category_name = str(row[1]).strip() if row[1] else False

            if not product_name:
                continue

            product = self.search([
                ('name', '=', product_name)
            ], limit=1)

            if not product:
                product_not_found.append(product_name)
                continue

            if not category_name:
                continue

            pos_category = self.env['pos.category'].search([
                ('name', '=', category_name)
            ], limit=1)

            if not pos_category:
                category_not_found.append(
                    f"{product_name} -> {category_name}"
                )
                continue

            product.write({
                'pos_categ_ids': [(6, 0, [pos_category.id])]
            })

            updated += 1

        _logger.info(
            "POS Category Mapping completed. Updated %s products.",
            updated
        )

        for product_name in product_not_found:
            _logger.warning(
                "Product not found: %s",
                product_name
            )

        for item in category_not_found:
            _logger.warning(
                "POS Category not found: %s",
                item
            )

    def _cron_product_vendor_mapping(self):

        attachment = self.env["ir.attachment"].search([
            ("name", "=", "product_vendor_mapping.xlsx")
        ], limit=1)

        if not attachment:
            _logger.info("Product vendor mapping file not found.")
            return

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(attachment.datas)),
            read_only=True,
            data_only=True,
        )

        sheet = workbook.active

        updated = 0
        product_not_found = []
        vendor_not_found = []

        SupplierInfo = self.env["product.supplierinfo"]
        Partner = self.env["res.partner"]

        for row in sheet.iter_rows(
                min_row=2,
                min_col=1,
                max_col=2,
                values_only=True,
        ):
            product_name = str(row[0]).strip() if row[0] else False
            vendor_name = str(row[1]).strip() if len(row) > 1 and row[1] else False

            if not product_name or not vendor_name:
                continue

            product = self.search([
                ("name", "=", product_name)
            ], limit=1)

            if not product:
                product_not_found.append(product_name)
                continue

            vendor = Partner.search([
                ("name", "=", vendor_name)
            ], limit=1)

            if not vendor:
                vendor_not_found.append(
                    f"{product_name} -> {vendor_name}"
                )
                continue

            existing_supplier = SupplierInfo.search([
                ("product_tmpl_id", "=", product.id),
                ("partner_id", "=", vendor.id),
            ], limit=1)

            if existing_supplier:
                continue

            SupplierInfo.create({
                "product_tmpl_id": product.id,
                "partner_id": vendor.id,
                "min_qty": 0,
            })

            updated += 1

        _logger.info(
            "Product Vendor Mapping completed. Updated %s products.",
            updated,
        )

        for product_name in product_not_found:
            _logger.warning("Product not found: %s", product_name)

        for item in vendor_not_found:
            _logger.warning("Vendor not found: %s", item)

    def _cron_update_supplierinfo_company(self):
        SupplierInfo = self.env["product.supplierinfo"].sudo()

        supplier_infos = SupplierInfo.search([
            ("company_id", "=", 1),
        ])

        count = len(supplier_infos)

        supplier_infos.write({
            "company_id": 2,
        })

        _logger.info(
            "SupplierInfo Company Update completed. "
            "Updated %s supplier records from company 1 to company 2.",
            count,
        )